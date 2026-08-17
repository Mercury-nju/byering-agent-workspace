"""
发票解析结果 → Excel 生成脚本

用途：读取 invoice_parsing 产出的 JSON 文件，生成格式化的 Excel 文件。
调用方式：python generate_invoice_excel.py --json-path <json路径> --output-path <输出路径>

⚠️ 本脚本由 invoice-retrieval 技能专用，禁止手动修改。
"""

import argparse
import json
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def main():
    parser = argparse.ArgumentParser(description="发票解析结果生成 Excel")
    parser.add_argument("--json-path", required=True, help="发票解析结果 JSON 文件路径")
    parser.add_argument("--output-path", required=True, help="Excel 输出路径")
    args = parser.parse_args()

    json_path = args.json_path
    output_path = args.output_path

    with open(json_path, "r", encoding="utf-8") as f:
        records = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票信息"

    # 判断是否存在非空备注内容（备注仅用于标记重复发票，无重复则不显示该列）
    has_remark = any(rec.get("备注", "").strip() for rec in records)

    # 表头（与观察结果字段一致，顺序固定：文件路径在销售方前）
    base_headers = ["序号", "发票号码", "开票日期", "价税合计", "税额", "不含税金额", "文件路径", "销售方", "购买方"]
    headers = base_headers + (["备注"] if has_remark else [])
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    # 填入数据（JSON 中的中文 key 直接映射）
    base_field_keys = ["发票号码", "开票日期", "价税合计", "税额", "不含税金额", "文件路径", "销售方", "购买方"]
    field_keys = base_field_keys + (["备注"] if has_remark else [])
    for idx, rec in enumerate(records, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        for col_offset, key in enumerate(field_keys, 2):
            val = rec.get(key, "")
            if key in ("价税合计", "税额", "不含税金额") and val:
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=idx + 1, column=col_offset, value=val)

    # 汇总行（价税合计=D列, 税额=E列, 不含税金额=F列）
    data_count = len(records)
    sum_row = data_count + 2
    ws.cell(row=sum_row, column=1, value="汇总")
    ws.cell(row=sum_row, column=1).font = Font(bold=True)
    ws.cell(row=sum_row, column=2, value=f"共 {data_count} 张发票")
    for col_idx, col_letter in [(4, "D"), (5, "E"), (6, "F")]:
        ws.cell(row=sum_row, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{sum_row - 1})")
        ws.cell(row=sum_row, column=col_idx).font = Font(bold=True)

    # 自动调整列宽
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    wb.save(output_path)
    print(f"Excel 已保存至: {output_path}")


if __name__ == "__main__":
    main()
