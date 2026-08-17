#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""子表探测与合计行写入工具。

本模块提供对同一 Sheet 内包含多个并排/上下排列子表的 Excel 文件进行：
1. 自动探测所有子表边界（先横向切分、后纵向切分）
2. 在每个子表末尾追加"合计"行（数值列写 =SUM 公式）

使用方式：
    from subtable_utils import add_total_rows_to_file
    output_path = add_total_rows_to_file("输入.xlsx")

或分步调用：
    from subtable_utils import detect_subtables, write_total_rows
    from openpyxl import load_workbook
    wb = load_workbook("输入.xlsx")
    ws = wb.active
    result = detect_subtables(ws)
    write_total_rows(ws, result)
    wb.save("输出.xlsx")
"""

import os
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def detect_subtables(ws: Worksheet) -> list[dict[str, Any]]:
    """探测 worksheet 中所有子表的边界。

    采用 8 连通域标记算法，能够完美识别任意错位并排的子表，
    只要子表之间至少有一行或一列的空白间隙。

    :param ws: openpyxl 的 Worksheet 对象
    :return: 子表列表，每项为 {"bounds": (r0, c0, r1, c1), "header": [...]}
    """
    max_row = ws.max_row
    max_col = ws.max_column

    # 构建布尔矩阵以加速访问
    grid = [[False] * (max_col + 2) for _ in range(max_row + 2)]
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                grid[r][c] = True

    visited = [[False] * (max_col + 2) for _ in range(max_row + 2)]
    result: list[dict[str, Any]] = []

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if grid[r][c] and not visited[r][c]:
                # BFS 寻找连通域
                queue = [(r, c)]
                visited[r][c] = True
                min_r, max_r = r, r
                min_c, max_c = c, c

                while queue:
                    curr_r, curr_c = queue.pop(0)
                    min_r = min(min_r, curr_r)
                    max_r = max(max_r, curr_r)
                    min_c = min(min_c, curr_c)
                    max_c = max(max_c, curr_c)

                    # 8 连通域
                    for dr in [-1, 0, 1]:
                        for dc in [-1, 0, 1]:
                            if dr == 0 and dc == 0:
                                continue
                            nr, nc = curr_r + dr, curr_c + dc
                            if 1 <= nr <= max_row and 1 <= nc <= max_col:
                                if grid[nr][nc] and not visited[nr][nc]:
                                    visited[nr][nc] = True
                                    queue.append((nr, nc))

                # 顶部修剪：去除可能紧挨着表头的单行大标题
                actual_min_r = min_r
                for row_idx in range(min_r, max_r + 1):
                    non_empty_count = sum(1 for col_idx in range(min_c, max_c + 1) if grid[row_idx][col_idx])
                    if non_empty_count >= 2:  # 真正的表头通常至少有2列
                        actual_min_r = row_idx
                        break

                # 过滤掉太小的连通域（如孤立的文本块）
                if max_r - actual_min_r + 1 >= 2 and max_c - min_c + 1 >= 2:
                    header = [ws.cell(row=actual_min_r, column=col).value for col in range(min_c, max_c + 1)]
                    result.append({
                        "bounds": (actual_min_r, min_c, max_r, max_c),
                        "header": header
                    })

    # 按位置排序：从上到下，从左到右
    result.sort(key=lambda x: (x["bounds"][0], x["bounds"][1]))
    return result


def write_total_rows(ws: Worksheet, result: list[dict[str, Any]]) -> None:
    """为每个子表在其末尾写入合计行。

    数值列写 Excel 原生 =SUM(...) 公式，非数值列留空。
    "合计"标签写在该子表的第一列（c0），不是 A 列。

    :param ws: openpyxl 的 Worksheet 对象
    :param result: detect_subtables 返回的子表列表
    """
    for blk in result:
        r0, c0, r1, c1 = blk["bounds"]
        total_row = r1 + 1

        # "合计"标签写在该子表的第一列
        ws.cell(row=total_row, column=c0, value="合计")

        for col in range(c0 + 1, c1 + 1):
            # 判定该列是否为"数值列"：遍历数据区（跳过表头行 r0）
            values = [
                ws.cell(row=r, column=col).value
                for r in range(r0 + 1, r1 + 1)
            ]

            has_numeric = False
            all_valid = True
            for v in values:
                if v is None or v == "":
                    continue
                if isinstance(v, (int, float)):
                    has_numeric = True
                elif isinstance(v, str) and v.startswith("="):
                    has_numeric = True
                else:
                    all_valid = False
                    break

            if has_numeric and all_valid:
                letter = get_column_letter(col)
                ws.cell(
                    row=total_row,
                    column=col,
                    value=f"=SUM({letter}{r0 + 1}:{letter}{r1})",
                )
            # 文本列留空，不写任何公式


def add_total_rows_to_file(
    file_path: str,
    output_path: str | None = None,
    sheet_name: str | None = None,
) -> str:
    """一站式入口：读取 Excel → 探测子表 → 写入合计行 → 另存副本。

    :param file_path: 输入 Excel 文件路径
    :param output_path: 输出文件路径，默认为 {原文件名}_带合计.xlsx
    :param sheet_name: 指定 Sheet 名，默认使用活动 Sheet
    :return: 输出文件的绝对路径
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 探测子表
    result = detect_subtables(ws)
    print(f"探测到 {len(result)} 个子表:")
    for i, blk in enumerate(result, 1):
        print(f"  子表{i}: bounds={blk['bounds']}, header={blk['header'][:5]}")

    # 写入合计行
    write_total_rows(ws, result)

    # 确定输出路径（禁止原地覆盖）
    if output_path is None:
        base, ext = os.path.splitext(file_path)
        output_path = f"{base}_带合计{ext}"

    wb.save(output_path)
    print(f"已保存到: {output_path}")
    return output_path


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法: python subtable_utils.py <Excel文件路径> [输出路径]")
        sys.exit(1)

    input_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else None
    add_total_rows_to_file(input_file, out_file)
